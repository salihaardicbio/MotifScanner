"""
report.py

Export motif scanning results.

Supported formats
-----------------
- CSV
- JSON
- Summary text

Author:
    Saliha Ardıç
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import WindowResult


class ReportWriter:
    """
    Export scan results to different file formats.
    """

    ##################################################################

    def __init__(self):

        pass

    ##################################################################

    @staticmethod
    def write_csv(
        results: list[WindowResult],
        output_file: str | Path
    ):
        """
        Export results to CSV.
        """

        output_file = Path(output_file)

        with open(
            output_file,
            "w",
            newline="",
            encoding="utf-8"
        ) as handle:

            writer = csv.writer(handle)

            writer.writerow([

                "Rank",

                "Sequence_ID",

                "Description",

                "Start",

                "End",

                "Peptide",

                "Total_Score",

                "Average_Score"

            ])

            for rank, result in enumerate(

                results,

                start=1

            ):

                writer.writerow([

                    rank,

                    result.sequence_id,

                    result.description,

                    result.start,

                    result.end,

                    result.peptide,

                    round(
                        result.total_score,
                        6
                    ),

                    round(
                        result.average_score,
                        6
                    )

                ])

    ##################################################################

    @staticmethod
    def write_json(
        results: list[WindowResult],
        output_file: str | Path
    ):
        """
        Export complete scan results to JSON.
        """

        output_file = Path(output_file)

        data = {

            "created":

                datetime.now().isoformat(),

            "hits": []

        }

        ##############################################################

        for rank, result in enumerate(

            results,

            start=1

        ):

            hit = {

                "rank":

                    rank,

                "sequence_id":

                    result.sequence_id,

                "description":

                    result.description,

                "start":

                    result.start,

                "end":

                    result.end,

                "peptide":

                    result.peptide,

                "total_score":

                    result.total_score,

                "average_score":

                    result.average_score,

                "positions": []

            }

            ##########################################################

            for position in result.position_scores:

                hit["positions"].append({

                    "position":

                        position.position,

                    "query":

                        position.query_residue,

                    "blosum":

                        position.blosum,

                    "property":

                        position.property,

                    "exact":

                        position.exact,

                    "conservation":

                        position.conservation,

                    "final":

                        position.final

                })

            ##########################################################

            data["hits"].append(
                hit
            )

        ##############################################################

        with open(
            output_file,
            "w",
            encoding="utf-8"
        ) as handle:

            json.dump(

                data,

                handle,

                indent=4

            )
    ##################################################################

    @staticmethod
    def write_summary(
        summary: dict,
        output_file: str | Path
    ):
        """
        Write scan summary to a text file.

        Parameters
        ----------
        summary
            Dictionary returned by Scanner.summary().

        output_file
            Output filename.
        """

        output_file = Path(output_file)

        with open(
            output_file,
            "w",
            encoding="utf-8"
        ) as handle:

            handle.write(
                "MotifScanner Summary Report\n"
            )

            handle.write(
                "=" * 40 + "\n\n"
            )

            handle.write(
                f"Proteins scanned : {summary['proteins']}\n"
            )

            handle.write(
                f"Windows scored   : {summary['windows']}\n"
            )

            handle.write(
                f"Hits retained    : {summary['hits']}\n"
            )

            handle.write(
                f"Best score       : "
                f"{summary['best_score']:.6f}\n"
            )

            handle.write(
                f"Elapsed time (s) : "
                f"{summary['elapsed_time']:.2f}\n"
            )

    ##################################################################

    @staticmethod
    def write_top_hits(
        results: list[WindowResult],
        output_file: str | Path,
        n: int = 20
    ):
        """
        Write the top N hits as a text report.
        """

        output_file = Path(output_file)

        with open(
            output_file,
            "w",
            encoding="utf-8"
        ) as handle:

            handle.write(
                f"Top {n} Motif Hits\n"
            )

            handle.write(
                "=" * 70 + "\n\n"
            )

            for rank, result in enumerate(
                results[:n],
                start=1
            ):

                handle.write(
                    f"Rank           : {rank}\n"
                )

                handle.write(
                    f"Protein        : {result.sequence_id}\n"
                )

                handle.write(
                    f"Description    : {result.description}\n"
                )

                handle.write(
                    f"Location       : "
                    f"{result.start}-{result.end}\n"
                )

                handle.write(
                    f"Peptide        : "
                    f"{result.peptide}\n"
                )

                handle.write(
                    f"Total Score    : "
                    f"{result.total_score:.6f}\n"
                )

                handle.write(
                    f"Average Score  : "
                    f"{result.average_score:.6f}\n"
                )

                handle.write(
                    "-" * 70 + "\n"
                )

    ##################################################################

    @staticmethod
    def write_position_details(
        results: list[WindowResult],
        output_file: str | Path
    ):
        """
        Export detailed position-by-position scores.

        One row corresponds to one motif position.
        """

        output_file = Path(output_file)

        with open(
            output_file,
            "w",
            newline="",
            encoding="utf-8"
        ) as handle:

            writer = csv.writer(handle)

            writer.writerow([

                "Sequence_ID",

                "Start",

                "Position",

                "Residue",

                "BLOSUM",

                "Property",

                "Exact",

                "Conservation",

                "Final"

            ])

            ##########################################################

            for result in results:

                for score in result.position_scores:

                    writer.writerow([

                        result.sequence_id,

                        result.start,

                        score.position,

                        score.query_residue,

                        round(score.blosum, 6),

                        round(score.property, 6),

                        score.exact,

                        round(score.conservation, 6),

                        round(score.final, 6)

                    ])
    ##################################################################

    @staticmethod
    def write_all(
        results: list[WindowResult],
        summary: dict,
        output_directory: str | Path,
        prefix: str = "motif_scan"
    ):
        """
        Write all report files.

        Parameters
        ----------
        results
            List of WindowResult objects.

        summary
            Summary dictionary returned by Scanner.summary().

        output_directory
            Destination directory.

        prefix
            Prefix for output filenames.

        Returns
        -------
        dict
            Dictionary containing generated file paths.
        """

        output_directory = Path(output_directory)

        output_directory.mkdir(
            parents=True,
            exist_ok=True
        )

        csv_file = output_directory / f"{prefix}_results.csv"

        json_file = output_directory / f"{prefix}_results.json"

        summary_file = output_directory / f"{prefix}_summary.txt"

        details_file = output_directory / (
            f"{prefix}_position_details.csv"
        )

        top_hits_file = output_directory / (
            f"{prefix}_top_hits.txt"
        )

        ##############################################################

        ReportWriter.write_csv(
            results,
            csv_file
        )

        ReportWriter.write_json(
            results,
            json_file
        )

        ReportWriter.write_summary(
            summary,
            summary_file
        )

        ReportWriter.write_position_details(
            results,
            details_file
        )

        ReportWriter.write_top_hits(
            results,
            top_hits_file
        )

        ##############################################################

        return {

            "csv": csv_file,

            "json": json_file,

            "summary": summary_file,

            "position_details": details_file,

            "top_hits": top_hits_file

        }

    ##################################################################

    @staticmethod
    def write_multi_motif_excel(
        rows: list[dict],
        motif_names: list[str],
        output_file: str | Path,
    ):
        """
        Writes a multi-motif comparison report to an Excel file.

        One row per protein. For every motif there are five columns
        (that motif's best-scoring window in this protein: start,
        end, peptide, total score, average score), followed by two
        combined columns: the product of every motif's best average
        score, and the product of every motif's best total score.

        Motifs may have different lengths, since each motif's best
        window is found independently.

        Parameters
        ----------
        rows
            List of dicts as produced by
            MultiMotifScanner.scan_records() - each containing
            sequence_id, description, and per motif
            "<motif>_start", "<motif>_end", "<motif>_peptide",
            "<motif>_total", "<motif>_average", plus
            multiplied_average and multiplied_total.

        motif_names
            Motif names in the order the columns should appear.

        output_file
            Destination .xlsx path.
        """

        output_file = Path(output_file)

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Multi-Motif Results"

        headers = [
            "Sequence_ID",
            "Description",
        ]

        for name in motif_names:

            headers.append(f"{name} - Start")
            headers.append(f"{name} - End")
            headers.append(f"{name} - Peptide")
            headers.append(f"{name} - Total_Score")
            headers.append(f"{name} - Average_Score")

        headers.append("Multiplied_Average_Score")
        headers.append("Multiplied_Total_Score")

        sheet.append(headers)

        for cell in sheet[1]:

            cell.font = Font(bold=True)

        for row in rows:

            values = [
                row["sequence_id"],
                row["description"],
            ]

            for name in motif_names:

                values.append(row[f"{name}_start"])
                values.append(row[f"{name}_end"])
                values.append(row[f"{name}_peptide"])
                values.append(round(row[f"{name}_total"], 6))
                values.append(round(row[f"{name}_average"], 6))

            values.append(round(row["multiplied_average"], 10))
            values.append(round(row["multiplied_total"], 6))

            sheet.append(values)

        ##############################################################
        # Auto-size columns.
        ##############################################################

        for column_index, header in enumerate(headers, start=1):

            column_letter = get_column_letter(column_index)

            longest = len(str(header))

            for row_cells in sheet.iter_rows(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
            ):

                value = row_cells[0].value

                if value is not None:
                    longest = max(longest, len(str(value)))

            sheet.column_dimensions[column_letter].width = min(
                max(longest + 2, 10),
                40,
            )

        sheet.freeze_panes = "A2"

        workbook.save(output_file)

    ##################################################################

    @staticmethod
    def print_report_locations(files: dict):
        """
        Print generated report locations.

        Parameters
        ----------
        files
            Dictionary returned by write_all().
        """

        print()

        print("=" * 60)

        print("Generated Report Files")

        print("=" * 60)

        for name, path in files.items():

            print(f"{name:20s} : {path}")

        print("=" * 60)


##########################################################################
# End of file
##########################################################################